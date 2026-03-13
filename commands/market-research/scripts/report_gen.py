#!/usr/bin/env python3
"""
report_gen.py - 시장조사 보고서 생성 (CSV + Markdown + Excel)

사용법:
    python3 report_gen.py --input data/analyzed.jsonl --format csv --output data/reports/
    python3 report_gen.py --input data/analyzed.jsonl --format markdown --output data/reports/
    python3 report_gen.py --input data/analyzed.jsonl --format excel --output data/reports/
    python3 report_gen.py --input data/analyzed.jsonl --format all --output data/reports/
"""

import argparse
import csv
import json
import sys
from datetime import datetime, date
from pathlib import Path


# ── CSV 출력 컬럼 정의 ────────────────────────────────────────────────
# 실무에서 가장 자주 조회하는 필드 순서로 정렬

CSV_COLUMNS = [
    # 기본 식별 정보
    ("플랫폼",              "platform"),
    ("상품유형",            "product_type"),
    ("검색키워드",          "search_keyword"),
    ("검색순위",            "search_rank"),
    ("광고여부",            "is_sponsored"),

    # 상품 정보
    ("상품명",              "product_name"),
    ("브랜드",              "brand_name"),
    ("제조사",              "manufacturer"),
    ("제조국",              "country_of_origin"),
    ("상품URL",             "product_url"),

    # 가격 정보 (핵심)
    ("정가",                "original_price"),
    ("판매가",              "sale_price"),
    ("할인율",              "discount_rate"),
    ("단위가격",            "unit_price"),
    ("단위기준",            "unit_price_basis"),
    ("쿠폰적용가",          "coupon_price"),
    ("최대할인가",          "max_discount_price"),
    ("가격대구간",          "price_tier"),
    ("가격백분위",          "price_percentile"),

    # 판매자 정보
    ("판매자명",            "seller_name"),
    ("판매자유형",          "seller_type"),
    ("판매자평점",          "seller_rating"),
    ("로켓배송",            "is_rocket_delivery"),
    ("공식브랜드스토어",    "is_brand_store"),

    # 스펙 (공통)
    ("장갑소재",            "glove_material"),
    ("장갑사이즈",          "glove_size"),
    ("장갑수량",            "glove_qty_per_pack"),
    ("분말여부",            "glove_has_powder"),
    ("백소재",              "bag_material"),
    ("백사이즈",            "bag_size"),
    ("백수량",              "bag_qty_per_pack"),
    ("지퍼유형",            "zipper_type"),
    ("냉동가능",            "freezer_safe"),
    ("전자레인지가능",      "microwave_safe"),
    ("식품용인증",          "food_grade_certified"),
    ("식품안전인증",        "food_safety_certified"),
    ("세트품목수",          "set_item_count"),

    # 판매·인기 지표
    ("누적판매량",          "total_sales_count"),
    ("최근30일판매",        "recent_sales_30d"),
    ("찜수",                "wishlist_count"),
    ("베스트셀러순위",      "bestseller_rank"),
    ("베스트셀러카테고리",  "bestseller_category"),
    ("재구매율",            "repurchase_rate_displayed"),

    # 리뷰
    ("평균평점",            "review_rating_avg"),
    ("리뷰수",              "review_count_total"),
    ("포토리뷰수",          "photo_review_count"),
    ("평점백분위",          "rating_percentile"),
    ("긍정키워드",          "positive_keywords"),
    ("부정키워드",          "negative_keywords"),
    ("불만카테고리",        "complaint_categories"),

    # 배송 정보
    ("배송유형",            "delivery_type"),
    ("무료배송",            "is_free_shipping"),
    ("배송비",              "shipping_cost"),
    ("무료배송최소주문",    "free_shipping_min_order"),

    # 프로모션
    ("행사여부",            "is_on_promotion"),
    ("행사유형",            "promotion_types"),
    ("타임딜",              "is_timedeal"),
    ("1+1",                 "is_one_plus_one"),
    ("쿠폰가능",            "coupon_available"),
    ("멤버십할인",          "membership_exclusive"),

    # 수집 메타
    ("수집일시",            "scraped_at"),
]


# ── CSV 생성 ──────────────────────────────────────────────────────────

def _format_cell(value) -> str:
    """CSV 셀 값 포맷팅."""
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'Y' if value else 'N'
    if isinstance(value, float):
        return f"{value:.2f}"
    if isinstance(value, list):
        return ' | '.join(str(v) for v in value) if value else ''
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def generate_csv(records: list[dict], output_path: Path) -> Path:
    """수집 데이터를 CSV로 저장. 날짜별 파일명 자동 생성."""
    today = date.today().strftime('%Y%m%d')
    csv_path = output_path / f"market_research_{today}.csv"

    headers = [col_name for col_name, _ in CSV_COLUMNS]
    field_keys = [field_key for _, field_key in CSV_COLUMNS]

    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for r in records:
            row = [_format_cell(r.get(key)) for key in field_keys]
            writer.writerow(row)

    return csv_path


# ── Markdown 리포트 생성 ──────────────────────────────────────────────

def _pct_bar(value: float, width: int = 20) -> str:
    """텍스트 프로그레스 바."""
    filled = int(value * width)
    return '█' * filled + '░' * (width - filled)


def generate_markdown(
    records: list[dict],
    output_path: Path,
    title: str = "주간 시장조사 리포트",
) -> Path:
    """수집 데이터를 Markdown 주간 리포트로 저장."""
    today = date.today()
    md_path = output_path / f"market_report_{today.strftime('%Y%m%d')}.md"

    lines = [
        f"# {title}",
        f"",
        f"**수집 기간**: {today.strftime('%Y년 %m월 %d일')}  ",
        f"**총 수집 상품**: {len(records):,}개  ",
        f"**생성 일시**: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"",
        "---",
        "",
    ]

    # 플랫폼별 요약
    platforms = sorted(set(r.get('platform', '') for r in records))
    product_types = ['위생장갑', '위생백', '지퍼백', '위생용품세트']

    for pt in product_types:
        subset = [r for r in records if r.get('product_type') == pt]
        if not subset:
            continue

        lines += [f"## {pt}", ""]

        for platform in platforms:
            group = [r for r in subset if r.get('platform') == platform]
            if not group:
                continue

            prices = [r['sale_price'] for r in group if r.get('sale_price', 0) > 0]
            unit_prices = [r['unit_price'] for r in group
                           if r.get('unit_price') and r['unit_price'] > 0]
            ratings = [r['review_rating_avg'] for r in group
                       if r.get('review_rating_avg', 0) > 0]
            disc_rates = [r['discount_rate'] for r in group
                          if r.get('discount_rate', 0) > 0]
            free_ship = sum(1 for r in group if r.get('is_free_shipping'))
            rocket = sum(1 for r in group if r.get('is_rocket_delivery'))
            sponsored = sum(1 for r in group if r.get('is_sponsored'))

            platform_label = {'coupang': '쿠팡', 'naver': '네이버쇼핑',
                               'gmarket': 'G마켓', '11st': '11번가'}.get(platform, platform)

            lines += [
                f"### {platform_label} ({len(group):,}개 상품)",
                "",
                "| 지표 | 값 |",
                "|------|----|",
                f"| 평균 판매가 | {int(sum(prices)/len(prices)):,}원 |" if prices else "| 평균 판매가 | - |",
                f"| 중앙 판매가 | {int(sorted(prices)[len(prices)//2]):,}원 |" if prices else "| 중앙 판매가 | - |",
                f"| 최저가 | {min(prices):,}원 |" if prices else "| 최저가 | - |",
                f"| 최고가 | {max(prices):,}원 |" if prices else "| 최고가 | - |",
                f"| 평균 단위가격 | {sum(unit_prices)/len(unit_prices):.1f}원/{group[0].get('unit_price_basis','개당')} |" if unit_prices else "| 평균 단위가격 | - |",
                f"| 평균 할인율 | {sum(disc_rates)/len(disc_rates)*100:.1f}% |" if disc_rates else "| 평균 할인율 | - |",
                f"| 평균 평점 | ★ {sum(ratings)/len(ratings):.2f} |" if ratings else "| 평균 평점 | - |",
                f"| 무료배송 비율 | {free_ship/len(group)*100:.0f}% ({free_ship}/{len(group)}) |",
                f"| 로켓배송 비율 | {rocket/len(group)*100:.0f}% ({rocket}/{len(group)}) |",
                f"| 광고 상품 비율 | {sponsored/len(group)*100:.0f}% ({sponsored}/{len(group)}) |",
                "",
            ]

            # 가격대 구간 분포
            tier_counts = {'저가': 0, '중가': 0, '고가': 0}
            for r in group:
                t = r.get('price_tier')
                if t in tier_counts:
                    tier_counts[t] += 1
            total_tier = sum(tier_counts.values())
            if total_tier > 0:
                lines += ["**가격대 구간 분포**", ""]
                lines += ["```"]
                for tier, cnt in tier_counts.items():
                    ratio = cnt / total_tier
                    lines.append(f"{tier:4s} {_pct_bar(ratio, 15)} {cnt:3d}개 ({ratio*100:.0f}%)")
                lines += ["```", ""]

            # 상위 브랜드
            from collections import Counter
            brand_cnt = Counter(r.get('brand_name', '') for r in group if r.get('brand_name'))
            if brand_cnt:
                lines += ["**상위 브랜드 (판매 상품 수 기준)**", ""]
                lines += ["| 순위 | 브랜드 | 상품 수 | 점유율 |",
                          "|------|--------|---------|--------|"]
                for rank, (brand, cnt) in enumerate(brand_cnt.most_common(5), 1):
                    share = cnt / len(group) * 100
                    lines.append(f"| {rank} | {brand} | {cnt} | {share:.1f}% |")
                lines += [""]

            # TOP 5 상품 (판매가 낮은 순)
            top5 = sorted(group, key=lambda r: r.get('sale_price', 0))[:5]
            if top5:
                lines += ["**최저가 TOP 5 상품**", ""]
                lines += ["| 순위 | 상품명 | 판매가 | 단위가격 | 평점 | 리뷰수 |",
                          "|------|--------|--------|----------|------|--------|"]
                for rank, r in enumerate(top5, 1):
                    name = r.get('product_name', '')[:30] + ('...' if len(r.get('product_name','')) > 30 else '')
                    price = f"{r.get('sale_price',0):,}원"
                    unit = f"{r.get('unit_price',0):.1f}원/{r.get('unit_price_basis','개당')}" if r.get('unit_price') else '-'
                    rating = f"★{r.get('review_rating_avg',0):.1f}"
                    reviews = f"{r.get('review_count_total',0):,}"
                    lines.append(f"| {rank} | {name} | {price} | {unit} | {rating} | {reviews} |")
                lines += [""]

            # 자주 언급 긍정/부정 키워드
            pos_all: Counter = Counter()
            neg_all: Counter = Counter()
            for r in group:
                for kw in r.get('positive_keywords', []):
                    pos_all[kw] += 1
                for kw in r.get('negative_keywords', []):
                    neg_all[kw] += 1

            if pos_all or neg_all:
                lines += ["**리뷰 키워드 집계**", ""]
                if pos_all:
                    top_pos = ', '.join(f"`{w}`" for w, _ in pos_all.most_common(8))
                    lines.append(f"- **긍정**: {top_pos}")
                if neg_all:
                    top_neg = ', '.join(f"`{w}`" for w, _ in neg_all.most_common(8))
                    lines.append(f"- **부정**: {top_neg}")
                lines += [""]

        lines += ["---", ""]

    lines += [
        "## 수집 방법론",
        "",
        "- 수집 플랫폼: 쿠팡, 네이버쇼핑",
        "- 수집 도구: MCP exa (URL 수집) + jina-reader (페이지 파싱)",
        "- 수집 대상: 위생장갑·위생백·지퍼백·위생용품세트",
        "- 단위 가격: 상품명/스펙에서 수량 추출 후 판매가 나누기",
        "- 가격대 구간: 동일 품목·플랫폼 내 하위33%=저가, 33~67%=중가, 상위33%=고가",
        "",
    ]

    md_path.write_text('\n'.join(lines), encoding='utf-8')
    return md_path


# ── Excel 생성 ────────────────────────────────────────────────────────

def generate_excel(
    records: list[dict],
    output_path: Path,
    title: str = "주간 시장조사 리포트",
) -> Path:
    """수집 데이터를 Excel(.xlsx)로 저장.

    시트 구성:
      - 전체데이터: 원시 레코드 (CSV와 동일한 컬럼)
      - 품목별요약: 품목×플랫폼 집계 통계
      - 최저가TOP10: 품목별 최저가 상위 10개
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[오류] openpyxl 패키지가 필요합니다: pip install openpyxl", file=sys.stderr)
        raise

    today = date.today().strftime('%Y%m%d')
    xl_path = output_path / f"market_research_{today}.xlsx"

    wb = openpyxl.Workbook()

    # ── 스타일 상수 ──
    HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT   = Font(bold=True, color="FFFFFF", size=10)
    TITLE_FONT    = Font(bold=True, size=13)
    SUBHDR_FILL   = PatternFill("solid", fgColor="2E75B6")
    SUBHDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL      = PatternFill("solid", fgColor="EBF3FB")
    CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    THIN          = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    def _style_header_row(ws, row: int, col_count: int):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            cell.border = THIN

    def _style_data_row(ws, row: int, col_count: int, alt: bool = False):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row, column=c)
            if alt:
                cell.fill = ALT_FILL
            cell.alignment = LEFT
            cell.border = THIN

    # ── Sheet 1: 전체데이터 ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "전체데이터"

    headers   = [col_name for col_name, _ in CSV_COLUMNS]
    field_keys = [fk for _, fk in CSV_COLUMNS]

    # 타이틀 행
    ws1.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    title_cell = ws1["A1"]
    title_cell.value = f"{title}  ({date.today().strftime('%Y-%m-%d')})"
    title_cell.font  = TITLE_FONT
    title_cell.alignment = CENTER
    ws1.row_dimensions[1].height = 24

    # 헤더 행
    for ci, h in enumerate(headers, 1):
        ws1.cell(row=2, column=ci, value=h)
    _style_header_row(ws1, 2, len(headers))
    ws1.row_dimensions[2].height = 20

    # 데이터 행
    for ri, r in enumerate(records, 3):
        for ci, fk in enumerate(field_keys, 1):
            val = r.get(fk)
            if isinstance(val, bool):
                val = "Y" if val else "N"
            elif isinstance(val, list):
                val = " | ".join(str(v) for v in val) if val else ""
            elif isinstance(val, dict):
                val = json.dumps(val, ensure_ascii=False)
            ws1.cell(row=ri, column=ci, value=val)
        _style_data_row(ws1, ri, len(headers), alt=(ri % 2 == 0))

    # 열 너비 자동 조정 (최대 40)
    for ci, h in enumerate(headers, 1):
        col_letter = get_column_letter(ci)
        max_len = max(
            len(str(h)),
            *[len(str(r.get(fk) or "")) for r in records],
            1
        )
        ws1.column_dimensions[col_letter].width = min(max_len + 2, 40)

    # 틀 고정 (헤더 2행 고정)
    ws1.freeze_panes = "A3"

    # ── Sheet 2: 품목별요약 ──────────────────────────────────────────
    ws2 = wb.create_sheet("품목별요약")
    product_types = ['위생장갑', '위생백', '지퍼백', '위생용품세트']
    platforms     = sorted(set(r.get('platform', '') for r in records))

    summary_headers = [
        "상품유형", "플랫폼", "상품수",
        "평균판매가(원)", "최저가(원)", "최고가(원)", "중앙값(원)",
        "평균단위가격", "평균할인율(%)", "평균평점",
        "무료배송(%)", "로켓배송(%)", "광고비율(%)",
    ]
    for ci, h in enumerate(summary_headers, 1):
        ws2.cell(row=1, column=ci, value=h)
    _style_header_row(ws2, 1, len(summary_headers))

    row_idx = 2
    for pt in product_types:
        for pl in platforms:
            group = [r for r in records
                     if r.get('product_type') == pt and r.get('platform') == pl]
            if not group:
                continue
            prices      = [r['sale_price']         for r in group if r.get('sale_price', 0) > 0]
            unit_prices = [r['unit_price']          for r in group if r.get('unit_price') and r['unit_price'] > 0]
            ratings     = [r['review_rating_avg']   for r in group if r.get('review_rating_avg', 0) > 0]
            disc_rates  = [r['discount_rate']       for r in group if r.get('discount_rate', 0) > 0]
            free_ship   = sum(1 for r in group if r.get('is_free_shipping'))
            rocket      = sum(1 for r in group if r.get('is_rocket_delivery'))
            sponsored   = sum(1 for r in group if r.get('is_sponsored'))

            n = len(group)
            row_data = [
                pt, pl, n,
                int(sum(prices) / len(prices)) if prices else 0,
                min(prices) if prices else 0,
                max(prices) if prices else 0,
                int(sorted(prices)[len(prices) // 2]) if prices else 0,
                round(sum(unit_prices) / len(unit_prices), 2) if unit_prices else 0,
                round(sum(disc_rates) / len(disc_rates) * 100, 1) if disc_rates else 0,
                round(sum(ratings) / len(ratings), 2) if ratings else 0,
                round(free_ship / n * 100, 1),
                round(rocket / n * 100, 1),
                round(sponsored / n * 100, 1),
            ]
            for ci, val in enumerate(row_data, 1):
                ws2.cell(row=row_idx, column=ci, value=val)
            _style_data_row(ws2, row_idx, len(summary_headers), alt=(row_idx % 2 == 0))
            row_idx += 1

    for ci, h in enumerate(summary_headers, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = max(len(h) + 2, 14)
    ws2.freeze_panes = "A2"

    # ── Sheet 3: 최저가TOP10 ──────────────────────────────────────────
    ws3 = wb.create_sheet("최저가TOP10")
    top_headers = [
        "상품유형", "플랫폼", "순위", "상품명", "브랜드",
        "판매가(원)", "단위가격", "단위기준", "할인율(%)",
        "평점", "리뷰수", "무료배송", "로켓배송", "상품URL",
    ]
    for ci, h in enumerate(top_headers, 1):
        ws3.cell(row=1, column=ci, value=h)
    _style_header_row(ws3, 1, len(top_headers))

    row_idx = 2
    for pt in product_types:
        for pl in platforms:
            group = sorted(
                [r for r in records
                 if r.get('product_type') == pt and r.get('platform') == pl
                 and r.get('sale_price', 0) > 0],
                key=lambda r: r.get('sale_price', 0),
            )[:10]
            for rank, r in enumerate(group, 1):
                row_data = [
                    pt, pl, rank,
                    r.get('product_name', ''),
                    r.get('brand_name', ''),
                    r.get('sale_price', 0),
                    r.get('unit_price') or '',
                    r.get('unit_price_basis', ''),
                    round(r.get('discount_rate', 0) * 100, 1),
                    r.get('review_rating_avg', 0),
                    r.get('review_count_total', 0),
                    "Y" if r.get('is_free_shipping') else "N",
                    "Y" if r.get('is_rocket_delivery') else "N",
                    r.get('product_url', ''),
                ]
                for ci, val in enumerate(row_data, 1):
                    ws3.cell(row=row_idx, column=ci, value=val)
                _style_data_row(ws3, row_idx, len(top_headers), alt=(row_idx % 2 == 0))
                row_idx += 1

    for ci, h in enumerate(top_headers, 1):
        ws3.column_dimensions[get_column_letter(ci)].width = max(len(h) + 2, 12)
    ws3.freeze_panes = "A2"

    wb.save(xl_path)
    return xl_path


# ── CLI ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="시장조사 보고서 생성기")
    parser.add_argument("--input", default="data/analyzed.jsonl", help="입력 JSONL 경로")
    parser.add_argument("--output", default="data/reports", help="출력 디렉터리")
    parser.add_argument("--format", choices=["csv", "markdown", "excel", "all"], default="all")
    parser.add_argument("--title", default="주간 시장조사 리포트", help="보고서 제목")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    output_path.mkdir(parents=True, exist_ok=True)

    if not input_path.exists():
        print(f"[error] 파일 없음: {input_path}", file=sys.stderr)
        sys.exit(1)

    records = []
    with open(input_path, encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line:
                records.append(json.loads(line))

    print(f"[info] {len(records):,}개 레코드 로드", file=sys.stderr)

    if args.format in ("csv", "all"):
        csv_file = generate_csv(records, output_path)
        print(f"[csv]  {csv_file}")

    if args.format in ("markdown", "all"):
        md_file = generate_markdown(records, output_path, args.title)
        print(f"[md]   {md_file}")

    if args.format in ("excel", "all"):
        xl_file = generate_excel(records, output_path, args.title)
        print(f"[xlsx] {xl_file}")
